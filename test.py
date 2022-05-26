from audioop import add
from pickle import FALSE, TRUE
import openpyxl
from sympy import false

rb = openpyxl.load_workbook('./srkv.xlsx')
sheet = rb.active
addit = TRUE
i = 1
while sheet.cell(row = 1, column = i).value: i+=1
resNu = [0,1,2,3]
for j in range(1,len(resNu)+1):
    sheet.cell(row = j, column = i).value = resNu[j-1]
rb.save('./srkv.xlsx')