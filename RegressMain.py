from itertools import count
import numpy as np
from scipy import optimize
import random as rnd
import math
# import NonParametrEstmation as npe
from scipy.sparse.sputils import getdata
from RandomVariables import RandomVariables as RV
# from scipy.optimize.optimize import main
import MLE

from tkinter import filedialog as fd
from Errors import MyError
from tkinter import messagebox as mb

No = 0
сountMembers = [5,8,34,5, 9]
valuesFurie = [u"Дискретный", u"Быстрый", u"Прямоугольное окно", u"Окно Ханна", u"Окно Хемминга"]
valuesModel = [u"Морле", u"Мексиканская шляпа", u"DOG", u"LP", u"Фурье"]
valuesDistribution = [u"Нормальное распределение", u"Экспоненциальное распределение", u"Гамма-распределение", u"Бета-распределение"]

def fileData(fileName):
    with open(fileName) as fin:
        data = [np.array(line.split()).astype(np.float64) for line in fin]
        Y, X = np.split(data, [1])                                   
    return Y[0], X.T

def fileThets(fileName):
    file = open(fileName)
    thets = [float(line) for line in file.readline().split()]
    file.close()
    return thets

def withNoise(Y, noise, drop, chd, bigNoise):
    n = len(Y)
    dy = sum(yi for yi in Y)/n
    w = sum((yi-dy)**2 for yi in Y)/(n-1) 
    y = [yi for yi in Y]
    
    parametrs = [0,1,0.5, 1, 9,2]
    randomVariables = RV(len(y), parametrs)
    rv = randomVariables.getFunction(chd) 
    
    for i, rvi in zip(range(len(y)), rv):
        requency = rnd.uniform(0,1)
        if(drop == '' or requency >= drop):
            y[i] += rvi*noise#*math.sqrt(w)                 #!!!!!!!!!!!!!!!!!!!!!!
        else:
            y[i] += rvi*bigNoise#*math.sqrt(w)
    return y

def func(theta, x):                                                     #   ВХОД: ТЕТЫ, x - равный размер
    return sum(theta[i]*xi for i, xi in zip(range(len(theta)), x))

def createData(direct, fileName, countV, noise, drop,chd, bigNoise, i):
    thets = fileThets(fileName)
    countT = len(thets)
    # Y, X = fileData(direct+"/my.txt")
    # X = []
    # x = sorted([rnd.uniform(-1,1) for j in range(countV)])
    # for j in range(countT):
    #     X.append([xi**j for xi in x])
    # X=np.array(X).T
    # Y = Y1 = [func(thets, xi) for xi in X]
    # Y = withNoise(Y, noise, drop,chd, bigNoise)
    # # e = np.array([yi - y1 for yi, y1 in zip(Y, Y1)])
    # data = np.vstack([np.array(Y), X.T])
    resFileName = direct + "/" + "my"+str(i) +" " + valuesDistribution[chd] + " "+ " N=" + str(countV) + " T=" + str(countT) + " noise=" + str(noise)+ " drop="+ str(drop)+" BigNoise="+str(bigNoise)+".txt"
    # np.savetxt(resFileName, data)
    return fileData(resFileName)

def getSP(X, Y):
    m = X.T @ X
    if np.linalg.det(m) == 0: raise MyError("вырожденная")
    thets = np.linalg.inv(X.T @ X) @ X.T @ Y
    return thets

def getXY(X, Y):
    size = int(len(Y)*0.5)
    I = [int(i/size * (len(Y)-1)) for i in range(size)]
    x = np.array([X[i] for i in I])
    y = np.array([Y[i] for i in I])
    return x,y

def getVSM(models):
    res = []
    for i in models:
        res.append(valuesModel[i])
    res.append("МНК")
    return res

import openpyxl

def printRes(fileName, resT, resNu, VSM, resNu2, noise):
    file = open(fileName, "w", encoding="utf-8")
    for model, t, in zip(VSM, resT):
        file.write("Received for {}: {}\n".format(model,t))   
    for nu in resNu:
        file.write("{0:.4f}\n".format(nu))
    file.write("\n")
    for nu in resNu2:
        file.write("{0:.4f}\n".format(nu))
    file.close()

    rb = openpyxl.load_workbook('./srkv.xlsx')
    sheet = rb.active
    i = 1
    while sheet.cell(row = 1, column = i).value: i+=1
    sheet.cell(row = 1, column = i).value = noise
    for j in range(len(resNu)):
        sheet.cell(row = j+2, column = i).value = resNu[j]
    rb.save('./srkv.xlsx')

def culcNuTest(fileName, T):
    trueT = fileThets(fileName)
    return sum(abs(ti - tTi)/abs(tTi) for ti, tTi in zip(T, trueT))
    # Y, X = fileData("./my.txt")
    # Y2 = [func(T, xi) for xi in X]
    # return math.sqrt(sum((y-y2)**2 for y,y2 in zip(Y,Y2)))/len(Y)

def mediumListNu(fileName, T):
    nus = []
    for t in T:
        nus.append(culcNuTest(fileName, t))
    return sum(nu for nu in nus)/len(nus)

def mediumListT(T):
    T = np.array(T).T
    mas = []
    for Ti in T:
        mas.append(sum(at for at in Ti)/len(Ti))
    return mas

def culcRes (file, fileRes, Xglob, Yglob, countM, countS, noise):
    mle = MLE.MLE(countM)
    thets = []
    nus2 = []
    nus = []
    startT = []
    for s in range(countS):
        X = Xglob[s]
        Y = Yglob[s]
        startT.append(getSP(X, Y))
        # print(startT[s])
    # print()
    for chm in range(len(valuesModel)):
        athets = []
        for s in range(countS):    
            x,y = getXY(X, Y)
            st = startT[s]
            # st = [2.01, -4.01, 9.01]
            bnds = [(t-20, t+20) for t in st]
            
            res = optimize.minimize(mle.algFT, st, args=(X,Y,X,Y,chm), method='SLSQP' #method='Nelder-Mead')
                                    # options={'xtol': 1e-8, 'disp': True})
                                    , bounds=bnds
                                    )
            # print(res.x)
            athets.append(res.x)
        # print()
        thets.append(mediumListT(athets)) # теты
        # nus2.append(mediumListNuDisp(file, athets))
        nus.append(mediumListNu(file, athets))        #среднее отклонение
    thets.append(mediumListT(startT))
    # nus2.append(mediumListNuDisp(file,startT))
    nus.append(mediumListNu(file, startT))
    VSM = getVSM(range(len(valuesModel)))
    printRes(fileRes, thets, nus, VSM, nus2, noise)

def main():
    try:
        direct = fd.askdirectory()
        if direct == '': raise MyError("Не выбрана дирректория")
        file = fd.askopenfilename()
        if file == '': raise MyError("Не выбран файл с параметрами")
        countV = 100
        for noise in [0.05]:#i/100 for i in range(5,51,5)]:
            for drop in [0.2]:#i/100 for i in range(5,51,5)]:
                bigNoise = 0.5
                chd = 0
                Xglob = []
                Yglob = []
                CountS = 20
                for s in range(CountS):
                    Y, X = createData(direct, file,  countV, noise, drop,chd, bigNoise, s)
                    Xglob.append(X)
                    Yglob.append(Y)
                fileName = direct + "/" + valuesDistribution[chd] + " " + str(noise)+ " "+ str(drop)+" "+str(bigNoise) + ".txt"
                culcRes(file, fileName, Xglob, Yglob, сountMembers, CountS, noise)
    except MyError as er:
            print(er)


main()
mb.showinfo("Message", "закончили")